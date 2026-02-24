from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Q
from django.core.paginator import Paginator
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .models import Product, Category, Purchase, PurchaseItem, Supplier


SORT_FIELDS = {
    "name": "name",
    "part": "part_number",
    "brand": "brand",
    "category": "category__name",
    "qty": "quantity",
    "price": "selling_price",
    "cost": "cost_price",
}


@login_required
def product_list(request):
    products = Product.objects.select_related("category")

    # Search
    q = request.GET.get("q", "").strip()
    if q:
        products = products.filter(
            Q(name__icontains=q)
            | Q(sku__icontains=q)
            | Q(part_number__icontains=q)
            | Q(brand__icontains=q)
            | Q(location__icontains=q)
            | Q(vehicle_application__icontains=q)
            | Q(purpose__icontains=q)
            | Q(compatibilities__name__icontains=q)
            | Q(compatibilities__make__name__icontains=q)
        ).distinct()

    # Filters
    category_id = request.GET.get("category", "")
    if category_id:
        products = products.filter(category_id=category_id)

    brand_filter = request.GET.get("brand", "").strip()
    if brand_filter:
        products = products.filter(brand__iexact=brand_filter)

    stock_filter = request.GET.get("stock", "")
    if stock_filter == "low":
        products = products.filter(quantity__lte=Product.LOW_STOCK_THRESHOLD)
    elif stock_filter == "ok":
        products = products.filter(quantity__gt=Product.LOW_STOCK_THRESHOLD)

    # Sorting
    sort = request.GET.get("sort", "name")
    order = request.GET.get("order", "asc")
    if sort in SORT_FIELDS and order in ("asc", "desc"):
        field = SORT_FIELDS[sort]
        products = products.order_by(("-" if order == "desc" else "") + field)

    # Pagination
    per_page = min(int(request.GET.get("per_page", 25) or 25), 100)
    per_page = max(per_page, 10)
    paginator = Paginator(products, per_page)
    page_num = request.GET.get("page", 1)
    page = paginator.get_page(page_num)

    # Filter options for dropdowns
    categories = Category.objects.all().order_by("name")
    brands = (
        Product.objects.exclude(brand="")
        .exclude(brand__isnull=True)
        .values_list("brand", flat=True)
        .distinct()
        .order_by("brand")
    )

    return render(request, "inventory/product_list.html", {
        "products": page,
        "page": page,
        "categories": categories,
        "brands": brands,
        "search": q,
        "sort": sort,
        "order": order,
        "filter_category": category_id,
        "filter_brand": brand_filter,
        "filter_stock": stock_filter,
        "per_page": per_page,
    })


@login_required
def product_detail(request, pk):
    """Show detailed information for a single product."""
    product = get_object_or_404(Product, pk=pk)
    return render(request, "inventory/product_detail.html", {"product": product})


@login_required
def export_inventory(request):
    """Export products to Excel file. Honors current filters when applied."""
    products = Product.objects.select_related("category")

    # Apply same filters as product_list for consistency
    q = request.GET.get("q", "").strip()
    if q:
        products = products.filter(
            Q(name__icontains=q)
            | Q(sku__icontains=q)
            | Q(part_number__icontains=q)
            | Q(brand__icontains=q)
            | Q(location__icontains=q)
            | Q(vehicle_application__icontains=q)
            | Q(purpose__icontains=q)
            | Q(compatibilities__name__icontains=q)
            | Q(compatibilities__make__name__icontains=q)
        ).distinct()
    category_id = request.GET.get("category", "")
    if category_id:
        products = products.filter(category_id=category_id)
    brand_filter = request.GET.get("brand", "").strip()
    if brand_filter:
        products = products.filter(brand__iexact=brand_filter)
    stock_filter = request.GET.get("stock", "")
    if stock_filter == "low":
        products = products.filter(quantity__lte=Product.LOW_STOCK_THRESHOLD)
    elif stock_filter == "ok":
        products = products.filter(quantity__gt=Product.LOW_STOCK_THRESHOLD)

    products = products.order_by("name")

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = [
        "Part Number", "Name", "Brand", "Category", "Quantity",
        "Cost Price", "Selling Price", "Location", "Purpose", "Vehicle/Car Types"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=product.part_number or "")
        ws.cell(row=row, column=2, value=product.name)
        ws.cell(row=row, column=3, value=product.brand or "")
        ws.cell(row=row, column=4, value=product.category.name if product.category else "")
        ws.cell(row=row, column=5, value=product.quantity)
        ws.cell(row=row, column=6, value=float(product.cost_price))
        ws.cell(row=row, column=7, value=float(product.selling_price))
        ws.cell(row=row, column=8, value=product.location or "")
        ws.cell(row=row, column=9, value=product.purpose or "")
        ws.cell(row=row, column=10, value=product.vehicle_application or "")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="inventory_export.xlsx"'
    wb.save(response)
    return response


@login_required
def import_inventory(request):
    """Import products from uploaded Excel file."""
    if request.method != "POST":
        return render(request, "inventory/import_inventory.html")

    excel_file = request.FILES.get("file")
    if not excel_file:
        return render(request, "inventory/import_inventory.html", {
            "error": "Please select an Excel file to upload."
        })

    if not excel_file.name.endswith(".xlsx"):
        return render(request, "inventory/import_inventory.html", {
            "error": "Invalid file type. Please upload an .xlsx file."
        })

    try:
        wb = load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active
        first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        first_cell = str(first_row[0] or "").upper() if first_row else ""

        # Detect format: "PART NAME" in header = spare parts format (INVENTORY.xlsx)
        is_spare_parts_format = first_cell and "PART NAME" in first_cell

        created = 0
        errors = []
        start_row = 2

        for row_num, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start_row):
            if not row or not any(row):
                continue

            if is_spare_parts_format:
                # Spare parts format: A=PART NAME, B=PART NUMBER, E=QTY, F=INSIDE PART NUMBER,
                # G=MARK (brand), H=TYPE (category), I=PURPOSE, J=FOR WHAT CAR
                name = str(row[0] or "").strip() if row[0] is not None else ""
                part_num_col_b = row[1]
                qty = row[4] if len(row) > 4 else None
                inside_part = str(row[5] or "").strip() if len(row) > 5 and row[5] else ""
                brand = str(row[6] or "").strip() if len(row) > 6 and row[6] else ""
                type_name = str(row[7] or "").strip() if len(row) > 7 and row[7] else ""
                purpose = str(row[8] or "").strip() if len(row) > 8 and row[8] else ""
                car_used = str(row[9] or "").strip() if len(row) > 9 and row[9] else ""

                # Skip section header rows (e.g. B0, C0, D0) - only col A has value
                if name and not any([part_num_col_b, qty, inside_part, brand, type_name]):
                    continue

                if not name:
                    continue

                part_number = inside_part or str(part_num_col_b or "").strip()
                quantity = int(qty or 0) if qty is not None else 0
                category_name = type_name
                vehicle_application = car_used
                location = ""  # Physical storage - set manually
                selling_price = 0.0  # Not in spare parts file - set in Admin after import
            else:
                # Standard format: Part Number, Name, Brand, Category, Qty, Cost, Price, Location, Purpose, Vehicle
                part_number = str(row[0] or "").strip() if row[0] is not None else ""
                name = str(row[1] or "").strip() if row[1] is not None else ""
                if not name:
                    errors.append(f"Row {row_num}: Name is required.")
                    continue
                brand = str(row[2] or "").strip() if len(row) > 2 and row[2] else ""
                category_name = str(row[3] or "").strip() if len(row) > 3 and row[3] else ""
                quantity = int(row[4] or 0) if len(row) > 4 and row[4] is not None else 0
                selling_price = float(row[5] or 0) if len(row) > 5 and row[5] is not None else 0.0
                location = str(row[6] or "").strip() if len(row) > 6 and row[6] else ""
                purpose_val = str(row[7] or "").strip() if len(row) > 7 and row[7] else ""
                vehicle_app = str(row[8] or "").strip() if len(row) > 8 and row[8] else ""

            category = None
            if category_name:
                category, _ = Category.objects.get_or_create(name=category_name)

            if is_spare_parts_format:
                purpose_val = purpose
                vehicle_app = vehicle_application

            Product.objects.create(
                part_number=part_number[:50],
                name=name[:100],
                brand=brand[:50] if brand else "",
                category=category,
                quantity=max(0, quantity),
                cost_price=0.0,
                selling_price=max(0, selling_price),
                location=location[:150] if location else "",
                purpose=purpose_val[:150] if purpose_val else "",
                vehicle_application=vehicle_app[:500] if vehicle_app else "",
            )
            created += 1

        wb.close()

        return render(request, "inventory/import_inventory.html", {
            "success": f"Successfully imported {created} product(s).",
            "errors": errors if errors else None,
        })

    except Exception as e:
        return render(request, "inventory/import_inventory.html", {
            "error": f"Import failed: {str(e)}"
        })


@login_required
def purchase_list(request):
    """List all purchases with date filter."""
    from django.db.models import Sum

    purchases = Purchase.objects.prefetch_related("items").select_related("supplier").order_by("-date")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    if date_from:
        purchases = purchases.filter(date__date__gte=date_from)
    if date_to:
        purchases = purchases.filter(date__date__lte=date_to)
    total_bought = sum(p.total_amount for p in purchases)
    return render(request, "inventory/purchase_list.html", {
        "purchases": purchases,
        "date_from": date_from,
        "date_to": date_to,
        "total_bought": total_bought,
    })


@login_required
def purchase_add(request):
    """Add a new purchase (bought items)."""
    products = Product.objects.select_related("category").order_by("name")
    suppliers = Supplier.objects.order_by("name")

    if request.method == "POST":
        supplier_id = request.POST.get("supplier") or None
        notes = request.POST.get("notes", "").strip()
        invoice_number = request.POST.get("invoice_number", "").strip()
        supplier = Supplier.objects.get(pk=supplier_id) if supplier_id else None

        with transaction.atomic():
            purchase = Purchase.objects.create(supplier=supplier, notes=notes, invoice_number=invoice_number)
            total = 0
            for product in products:
                qty = int(request.POST.get(f"qty_{product.id}", 0) or 0)
                if qty > 0:
                    unit_cost = float(request.POST.get(f"cost_{product.id}", 0) or 0)
                    PurchaseItem.objects.create(
                        purchase=purchase,
                        product=product,
                        quantity=qty,
                        unit_cost=unit_cost,
                    )
                    total += qty * unit_cost
            purchase.total_amount = total
            purchase.save()
        return redirect("purchase_list")

    return render(request, "inventory/purchase_add.html", {
        "products": products,
        "suppliers": suppliers,
    })
